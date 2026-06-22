#!/usr/bin/env python3
"""Build Taijiyou regional reading report from WeChat exported workbooks."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ORG_GROUPS = {
    "高客": ["北京", "浙江", "广东", "江苏", "福建", "天津", "上海"],
    "双战略": ["山东", "湖北", "四川", "河南", "湖南", "河北", "辽宁", "陕西", "安徽", "重庆", "云南", "山西"],
    "基盘": ["江西", "广西", "内蒙古", "甘肃", "吉林", "贵州", "黑龙江", "新疆", "宁夏", "海南", "青海", "西藏"],
}

CITIES_WITHOUT_SEPARATE_DATA = ["青岛", "深圳", "大连", "厦门", "宁波"]


def parse_date(value: Any) -> dt.date | None:
    if value in ("", None) or pd.isna(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def parse_ratio(value: Any) -> float:
    if value in ("", None) or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        if not text:
            return 0.0
        if text.endswith("%"):
            return float(text[:-1].strip()) / 100.0
        number = float(text)
    return number / 100.0 if number > 1 else number


def extract_region_records(path: Path) -> list[dict[str, Any]]:
    df_raw = pd.read_excel(path, header=None)
    if len(df_raw) < 10:
        return []

    all_texts = df_raw.iloc[0:10].stack().dropna().astype(str).tolist()
    title = max(all_texts, key=len) if all_texts else "未知标题"
    title = title.replace("\n", "").replace("\r", "").strip()

    publish_date: dt.date | None = None
    for row_idx in range(len(df_raw) - 1):
        cell_val = str(df_raw.iloc[row_idx, 1])
        if "日期" in cell_val:
            publish_date = parse_date(df_raw.iloc[row_idx + 1, 1])
            break

    try:
        total_reads = float(df_raw.iloc[4, 2])
    except Exception:
        total_reads = 0.0

    start_row = -1
    for row_idx in range(len(df_raw)):
        if "省份/直辖市" in str(df_raw.iloc[row_idx, 1]):
            start_row = row_idx + 1
            break

    if start_row == -1:
        return []

    records: list[dict[str, Any]] = []
    for row_idx in range(start_row, len(df_raw)):
        region_value = df_raw.iloc[row_idx, 1]
        ratio_value = df_raw.iloc[row_idx, 2]
        if pd.isna(region_value):
            break
        region_name = str(region_value).strip()
        if not region_name or "总计" in region_name:
            break
        try:
            ratio = parse_ratio(ratio_value)
        except Exception:
            ratio = 0.0
        records.append(
            {
                "标题": title,
                "发布日期": publish_date.isoformat() if publish_date else "未知日期",
                "文章总阅读": int(round(total_reads)),
                "机构": region_name,
                "机构阅读人数": int(round(total_reads * ratio)),
                "来源文件": path.name,
            }
        )
    return records


def collect_region_records(folder: Path, output_filename: str, start: dt.date | None = None, end: dt.date | None = None) -> list[dict[str, Any]]:
    if not folder.exists() or not folder.is_dir():
        raise RuntimeError(f"地域表格文件夹不存在：{folder}")

    records: list[dict[str, Any]] = []
    for path in sorted(folder.iterdir(), key=lambda item: item.name):
        if path.name.startswith("~$"):
            continue
        if path.name == output_filename:
            continue
        if path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        try:
            file_records = extract_region_records(path)
        except Exception as exc:
            raise RuntimeError(f"读取地域表格失败：{path.name}；{exc}") from exc
        for record in file_records:
            publish_date = parse_date(record.get("发布日期"))
            if start and publish_date and publish_date < start:
                continue
            if end and publish_date and publish_date > end:
                continue
            records.append(record)
    return records


def province_read_total(province_summary: dict[str, int], target_name: str) -> int:
    return sum(value for name, value in province_summary.items() if target_name in name)


def sorted_group_stats(province_summary: dict[str, int], province_names: list[str]) -> list[dict[str, Any]]:
    stats = [{"名称": province, "阅读总数": province_read_total(province_summary, province)} for province in province_names]
    return sorted(stats, key=lambda item: item["阅读总数"], reverse=True)


def top3_text(stats: list[dict[str, Any]]) -> str:
    names = [item["名称"] for item in stats[:3] if item["阅读总数"] > 0]
    if len(names) >= 3:
        return f"{names[0]}、{names[1]}和{names[2]}"
    return "和".join(names) if names else "暂无"


def reads_display(total: int) -> str:
    if total < 10000:
        return f"{total}次"
    thousands = total // 1000
    remainder = total % 1000
    if remainder == 0:
        return f"{thousands / 10.0:.1f}万次"
    if remainder <= 499:
        return f"{thousands / 10.0:.1f}万余次"
    return f"近{(thousands + 1) / 10.0:.1f}万次"


def date_title_parts(df: pd.DataFrame) -> tuple[str, str]:
    dates = pd.to_datetime(df["发布日期"], errors="coerce").dropna()
    if dates.empty:
        return "未知时间", "本月"
    min_date = dates.min()
    max_date = dates.max()
    date_text = f"{min_date.year}.{min_date.month}.{min_date.day}-{max_date.month}.{max_date.day}"
    month_title = f"{min_date.year}年{min_date.month}月"
    return date_text, month_title


def build_region_report(folder: Path, output_path: Path, start: dt.date | None = None, end: dt.date | None = None) -> dict[str, Any]:
    records = collect_region_records(folder, output_path.name, start=start, end=end)
    if not records:
        raise RuntimeError("未提取到有效地域数据，请确认文件夹内有微信后台导出的地域 Excel。")

    df = pd.DataFrame(records)
    province_summary = df.groupby("机构")["机构阅读人数"].sum().astype(int).to_dict()

    group_stats = {name: sorted_group_stats(province_summary, provinces) for name, provinces in ORG_GROUPS.items()}
    group_totals = {name: sum(item["阅读总数"] for item in stats) for name, stats in group_stats.items()}
    total_system = sum(group_totals.values())
    date_text, month_title = date_title_parts(df)
    article_count = df["标题"].nunique()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet("统计汇总", 0)

        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)

        worksheet.merge_cells("B1:O1")
        worksheet["B1"] = f"{month_title}泰绩优公众号阅读量机构统计"
        worksheet["B1"].font = Font(size=16, bold=True)
        worksheet["B1"].alignment = align_center

        summary_text = (
            f"用好线上内容矩阵，赋能最大绩优。{date_text}，泰绩优公众号共发布赋能推文{article_count}篇，"
            f"累计阅览量{reads_display(total_system)}，其中高客阅览量前三：{top3_text(group_stats['高客'])}；"
            f"双战略阅览量前三：{top3_text(group_stats['双战略'])}；"
            f"基盘阅览量前三：{top3_text(group_stats['基盘'])}。"
        )
        worksheet.merge_cells("B2:O4")
        worksheet["B2"] = summary_text
        worksheet["B2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        worksheet["B2"].border = border

        sections = [("B", "高客", group_totals["高客"]), ("G", "双战略", group_totals["双战略"]), ("L", "基盘", group_totals["基盘"])]
        for start_col, label, subtotal in sections:
            start_idx = worksheet[f"{start_col}1"].column
            headers = ["类型", "机构", "阅读人次", "占比"]
            for offset, header in enumerate(headers):
                cell = worksheet.cell(row=5, column=start_idx + offset)
                cell.value = header
                cell.fill = header_fill
                cell.border = border
                cell.alignment = align_center
                cell.font = bold_font

            worksheet.merge_cells(start_row=6, start_column=start_idx, end_row=6, end_column=start_idx + 1)
            worksheet.cell(row=6, column=start_idx).value = "全系统"
            worksheet.cell(row=6, column=start_idx + 2).value = total_system
            for offset in range(4):
                worksheet.cell(row=6, column=start_idx + offset).border = border

            worksheet.merge_cells(start_row=7, start_column=start_idx, end_row=7, end_column=start_idx + 1)
            worksheet.cell(row=7, column=start_idx).value = f"{label}小计"
            worksheet.cell(row=7, column=start_idx + 2).value = subtotal
            ratio_cell = worksheet.cell(row=7, column=start_idx + 3)
            ratio_cell.value = subtotal / total_system if total_system else 0
            ratio_cell.number_format = "0.00%"
            for offset in range(4):
                worksheet.cell(row=7, column=start_idx + offset).border = border

        data_configs = [(group_stats["高客"], "B", group_totals["高客"], "高客"), (group_stats["双战略"], "G", group_totals["双战略"], "双战略"), (group_stats["基盘"], "L", group_totals["基盘"], "基盘")]
        for stats, column, subtotal, label in data_configs:
            start_idx = worksheet[f"{column}1"].column
            start_row = 8
            end_row = 19
            for index, item in enumerate(stats):
                row = start_row + index
                worksheet.cell(row=row, column=start_idx + 1).value = item["名称"]
                worksheet.cell(row=row, column=start_idx + 2).value = item["阅读总数"]
                ratio_cell = worksheet.cell(row=row, column=start_idx + 3)
                ratio_cell.value = item["阅读总数"] / subtotal if subtotal else 0
                ratio_cell.number_format = "0.00%"

            for row in range(start_row, end_row + 1):
                for offset in range(4):
                    worksheet.cell(row=row, column=start_idx + offset).border = border

            worksheet.merge_cells(start_row=start_row, start_column=start_idx, end_row=end_row, end_column=start_idx)
            merged_cell = worksheet.cell(row=start_row, column=start_idx)
            merged_cell.value = label
            merged_cell.alignment = align_center
            merged_cell.font = bold_font

        for index, city in enumerate(CITIES_WITHOUT_SEPARATE_DATA):
            cell = worksheet.cell(row=15 + index, column=3)
            cell.value = city
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

        worksheet.merge_cells("D15:E19")
        remark_cell = worksheet["D15"]
        remark_cell.value = "备注：公众号官方运营数据分析，阅读用户的地域分布按省份归属统计（北京除外），故青岛、深圳、大连、厦门和宁波暂无法单独统计至阅览人次"
        remark_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        remark_cell.font = Font(size=11)
        for row in range(15, 20):
            for column in range(4, 6):
                worksheet.cell(row=row, column=column).border = border

        column_widths = {"A": 3, "B": 11, "C": 12, "D": 12, "E": 13, "F": 3, "G": 11, "H": 12, "I": 12, "J": 13, "K": 3, "L": 11, "M": 12, "N": 12, "O": 13}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
        for row in range(1, 20):
            worksheet.row_dimensions[row].height = 25
        worksheet.row_dimensions[2].height = 42

        detail = df.copy()
        detail["sort_date"] = pd.to_datetime(detail["发布日期"], errors="coerce")
        detail = detail.sort_values(by=["sort_date", "标题", "机构"], ascending=[True, True, True]).drop(columns=["sort_date"])
        detail.to_excel(writer, sheet_name="数据明细", index=False)

    return {
        "records": len(records),
        "articles": int(article_count),
        "totalSystem": int(total_system),
        "output": str(output_path),
    }
