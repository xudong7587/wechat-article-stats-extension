#!/usr/bin/env python3
"""Local browser UI for WeChat Official Account analytics export."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import shutil
import subprocess
import sys
import threading
import time
import uuid
import webbrowser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import region_report

import wechat_stats


APP_TITLE = "公众号数据导出"
HOST = "127.0.0.1"

JOBS: dict[str, dict[str, Any]] = {}
JOBS_LOCK = threading.Lock()

FIELD_LABELS = {
    "source_type": "来源",
    "ref_date": "文章日期",
    "stat_date": "统计截至日期",
    "title": "文章标题",
    "content_url": "文章链接",
    "read_user": "阅读/播放数",
    "like_user": "点赞",
    "share_user": "转发",
    "zaikan_user": "喜欢",
    "comment_count": "留言/评论",
    "collection_user": "收藏",
    "read_finish_rate": "阅读完成率",
    "read_avg_activetime": "平均阅读时长(分钟)",
    "read_subscribe_user": "阅读后关注",
}

COMBINED_LABEL_OVERRIDES = {"like_user": "点赞/推荐"}

DEFAULT_SELECTED_FIELDS = [
    "title",
    "ref_date",
    "read_user",
    "like_user",
    "share_user",
    "zaikan_user",
    "comment_count",
    "source_type",
]

DEFAULT_GUI_SETTINGS = {
    "version": 2,
    "mode": "totaldetail",
    "rowMode": "latest",
    "selectedFields": DEFAULT_SELECTED_FIELDS[:],
}

VIDEO_RAW_HEADERS = [
    "视频描述",
    "视频ID",
    "发布时间",
    "完播率",
    "平均播放时长",
    "播放量",
    "推荐",
    "喜欢",
    "评论量",
    "分享量",
    "关注量",
    "转发聊天和朋友圈",
    "设为铃声",
    "设为状态",
    "设为朋友圈封面",
]


HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>公众号数据导出</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f8;
      --panel: #ffffff;
      --text: #1f2329;
      --muted: #646a73;
      --border: #dfe3e8;
      --accent: #13795b;
      --accent-dark: #0e654c;
      --danger: #b42318;
      --warning: #9a6700;
      --ok: #13795b;
      --soft: #eef6f3;
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      font-family: "Microsoft YaHei UI", "Microsoft YaHei", system-ui, sans-serif;
      color: var(--text);
      background: var(--bg);
    }

    .topbar {
      height: 64px;
      display: flex;
      align-items: center;
      justify-content: space-between;
      padding: 0 28px;
      background: var(--panel);
      border-bottom: 1px solid var(--border);
    }

    .brand {
      display: flex;
      align-items: center;
      gap: 12px;
      font-weight: 700;
      font-size: 20px;
    }

    .brand-mark {
      width: 32px;
      height: 32px;
      border-radius: 8px;
      display: grid;
      place-items: center;
      color: white;
      background: var(--accent);
      font-weight: 800;
    }

    .status-pill {
      display: inline-flex;
      align-items: center;
      gap: 8px;
      min-height: 32px;
      padding: 0 12px;
      border: 1px solid var(--border);
      border-radius: 999px;
      color: var(--muted);
      background: #fff;
      font-size: 13px;
    }

    .dot {
      width: 8px;
      height: 8px;
      border-radius: 50%;
      background: var(--warning);
    }

    .dot.ok { background: var(--ok); }
    .dot.bad { background: var(--danger); }

    main {
      max-width: 1180px;
      margin: 0 auto;
      padding: 26px 28px 40px;
      display: grid;
      grid-template-columns: minmax(360px, 460px) minmax(360px, 1fr);
      gap: 20px;
    }

    section {
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 22px;
    }

    h1, h2 {
      margin: 0;
      letter-spacing: 0;
    }

    h1 { font-size: 22px; }
    h2 { font-size: 16px; margin-bottom: 18px; }

    .subtle {
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
    }

    .form-grid {
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 16px;
    }

    .field {
      display: flex;
      flex-direction: column;
      gap: 7px;
      margin-bottom: 16px;
    }

    label {
      color: #30343b;
      font-size: 13px;
      font-weight: 650;
    }

    input, select {
      width: 100%;
      min-height: 40px;
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 8px 10px;
      color: var(--text);
      background: white;
      font: inherit;
      font-size: 14px;
    }

    input:focus, select:focus {
      outline: 2px solid rgba(19, 121, 91, .16);
      border-color: var(--accent);
    }

    .checkbox-row {
      display: flex;
      align-items: center;
      gap: 9px;
      min-height: 36px;
      color: var(--muted);
      font-size: 13px;
    }

    .checkbox-row input {
      width: 16px;
      height: 16px;
      min-height: 16px;
    }

    .actions {
      display: flex;
      gap: 10px;
      align-items: center;
      flex-wrap: wrap;
      margin-top: 8px;
    }

    button {
      min-height: 40px;
      border: 1px solid transparent;
      border-radius: 6px;
      padding: 0 16px;
      font: inherit;
      font-weight: 650;
      cursor: pointer;
    }

    .primary {
      background: var(--accent);
      color: #fff;
    }

    .primary:hover { background: var(--accent-dark); }

    .secondary {
      border-color: var(--border);
      background: #fff;
      color: var(--text);
    }

    button:disabled {
      cursor: not-allowed;
      opacity: .6;
    }

    .notice {
      border-radius: 8px;
      border: 1px solid #d9eadf;
      background: var(--soft);
      padding: 14px;
      color: #23483d;
      font-size: 13px;
      line-height: 1.6;
      margin-bottom: 16px;
    }

    .notice.bad {
      border-color: #f1c4bd;
      background: #fff3f1;
      color: var(--danger);
    }

    .metrics {
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 10px;
      margin-bottom: 16px;
    }

    .metric {
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 14px;
      background: #fbfcfc;
    }

    .metric strong {
      display: block;
      font-size: 22px;
      margin-bottom: 4px;
    }

    .metric span {
      color: var(--muted);
      font-size: 12px;
    }

    .log {
      height: 292px;
      overflow: auto;
      border: 1px solid var(--border);
      border-radius: 8px;
      background: #101418;
      color: #d5e4df;
      padding: 14px;
      font-family: Consolas, "Microsoft YaHei UI", monospace;
      font-size: 13px;
      line-height: 1.55;
      white-space: pre-wrap;
    }

    .files {
      margin-top: 16px;
      display: grid;
      gap: 8px;
    }

    .file-row {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 10px 12px;
      background: #fff;
      font-size: 13px;
    }

    .file-row code {
      color: var(--muted);
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }

    @media (max-width: 900px) {
      .topbar { padding: 0 18px; }
      main {
        grid-template-columns: 1fr;
        padding: 18px;
      }
      .form-grid, .metrics { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <header class="topbar">
    <div class="brand"><span class="brand-mark">微</span><span>公众号数据导出</span></div>
    <div id="configStatus" class="status-pill"><span class="dot"></span><span>正在检查配置</span></div>
  </header>

  <main>
    <section>
      <h1>选择文章日期范围</h1>
      <p class="subtle">默认查询昨天。微信官方建议上午 8 点后查询前一天数据。</p>

      <div id="notice" class="notice" style="margin-top:18px;">填写 AppID 和 AppSecret 后即可开始导出。</div>

      <form id="exportForm">
        <div class="form-grid">
          <div class="field">
            <label for="start">开始日期</label>
            <input id="start" name="start" type="date" required>
          </div>
          <div class="field">
            <label for="end">结束日期</label>
            <input id="end" name="end" type="date" required>
          </div>
        </div>

        <div class="field">
          <label for="mode">数据接口</label>
          <select id="mode" name="mode">
            <option value="totaldetail">新接口：发表内容详细数据（推荐）</option>
            <option value="legacy-summary">旧接口：图文群发每日数据</option>
            <option value="legacy-total">旧接口：图文群发总数据</option>
          </select>
        </div>

        <div class="field">
          <label for="out">导出目录</label>
          <input id="out" name="out" type="text" placeholder="exports">
        </div>

        <label class="checkbox-row">
          <input id="forceRefreshToken" name="forceRefreshToken" type="checkbox">
          重新获取调用凭证
        </label>

        <div class="actions">
          <button id="submitBtn" class="primary" type="submit">开始导出</button>
          <button id="openFolderBtn" class="secondary" type="button">打开导出文件夹</button>
        </div>
      </form>
    </section>

    <section>
      <h2>导出结果</h2>
      <div class="metrics">
        <div class="metric"><strong id="rowsCount">0</strong><span>主表行数</span></div>
        <div class="metric"><strong id="sourceCount">0</strong><span>阅读来源行数</span></div>
        <div class="metric"><strong id="jumpCount">0</strong><span>跳出位置行数</span></div>
      </div>
      <div id="log" class="log">等待开始...</div>
      <div id="files" class="files"></div>
    </section>
  </main>

  <script>
    const $ = (id) => document.getElementById(id);
    let activeJobId = null;
    let pollTimer = null;

    function setStatus(configured, text) {
      const box = $("configStatus");
      const dot = box.querySelector(".dot");
      dot.className = "dot " + (configured ? "ok" : "bad");
      box.querySelector("span:last-child").textContent = text;
    }

    function appendLog(text) {
      $("log").textContent = text || "等待开始...";
      $("log").scrollTop = $("log").scrollHeight;
    }

    function setBusy(busy) {
      $("submitBtn").disabled = busy;
      $("submitBtn").textContent = busy ? "正在导出..." : "开始导出";
    }

    async function api(path, options) {
      const response = await fetch(path, options);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "请求失败");
      return data;
    }

    async function loadStatus() {
      const data = await api("/api/status");
      $("start").value = data.yesterday;
      $("end").value = data.yesterday;
      $("end").max = data.yesterday;
      $("start").max = data.yesterday;
      $("out").value = data.exportDir;
      setStatus(data.configured, data.configured ? "配置已就绪" : "缺少 AppID/AppSecret");
      $("notice").className = data.configured ? "notice" : "notice bad";
      $("notice").textContent = data.configured
        ? "配置已就绪。选择日期范围后点击开始导出。"
        : "请先打开 .env 文件，填入 WECHAT_APPID 和 WECHAT_APPSECRET。";
    }

    async function pollJob() {
      if (!activeJobId) return;
      const data = await api("/api/job?id=" + encodeURIComponent(activeJobId));
      appendLog(data.log.join("\n"));
      $("rowsCount").textContent = data.rows || 0;
      $("sourceCount").textContent = data.sources || 0;
      $("jumpCount").textContent = data.jumps || 0;
      renderFiles(data.files || []);
      if (data.status === "done" || data.status === "error") {
        clearInterval(pollTimer);
        pollTimer = null;
        setBusy(false);
      }
    }

    function renderFiles(files) {
      const wrap = $("files");
      wrap.innerHTML = "";
      for (const file of files) {
        const row = document.createElement("div");
        row.className = "file-row";
        const code = document.createElement("code");
        code.textContent = file;
        row.appendChild(code);
        wrap.appendChild(row);
      }
    }

    $("exportForm").addEventListener("submit", async (event) => {
      event.preventDefault();
      setBusy(true);
      renderFiles([]);
      appendLog("准备开始...");
      $("rowsCount").textContent = "0";
      $("sourceCount").textContent = "0";
      $("jumpCount").textContent = "0";

      try {
        const payload = {
          start: $("start").value,
          end: $("end").value,
          mode: $("mode").value,
          out: $("out").value,
          forceRefreshToken: $("forceRefreshToken").checked
        };
        const data = await api("/api/export", {
          method: "POST",
          headers: {"Content-Type": "application/json"},
          body: JSON.stringify(payload)
        });
        activeJobId = data.jobId;
        pollTimer = setInterval(pollJob, 900);
        await pollJob();
      } catch (error) {
        setBusy(false);
        appendLog("失败：" + error.message);
      }
    });

    $("openFolderBtn").addEventListener("click", async () => {
      try {
        await api("/api/open-folder", {
          method: "POST",
          headers: {"Content-Type": "application/json"},
          body: JSON.stringify({out: $("out").value})
        });
      } catch (error) {
        appendLog("打开文件夹失败：" + error.message);
      }
    });

    loadStatus().catch((error) => appendLog("加载失败：" + error.message));
  </script>
</body>
</html>
"""


def today_yesterday() -> dt.date:
    return dt.date.today() - dt.timedelta(days=1)


def configured() -> bool:
    wechat_stats.load_dotenv(wechat_stats.repo_path(".env"))
    appid = os.environ.get("WECHAT_APPID", "").strip()
    secret = os.environ.get("WECHAT_APPSECRET", "").strip()
    return bool(appid and secret and not appid.startswith("你的") and not secret.startswith("你的"))


def resolve_export_dir(value: str | None) -> Path:
    if value:
        path = Path(value.strip())
    else:
        path = Path(os.environ.get("WECHAT_EXPORT_DIR", "exports"))
    if not path.is_absolute():
        path = wechat_stats.repo_path(str(path))
    return path.resolve()


def load_settings() -> dict[str, Any]:
    path = wechat_stats.repo_path(".gui_settings.json")
    if not path.exists():
        return dict(DEFAULT_GUI_SETTINGS)
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return dict(DEFAULT_GUI_SETTINGS)
    fields = data.get("selectedFields")
    if not isinstance(fields, list):
        fields = DEFAULT_SELECTED_FIELDS[:]
    fields = [str(field) for field in fields if str(field) in FIELD_LABELS]
    mode = str(data.get("mode") or DEFAULT_GUI_SETTINGS["mode"])
    if mode not in wechat_stats.ENDPOINTS:
        mode = DEFAULT_GUI_SETTINGS["mode"]
    row_mode = str(data.get("rowMode") or DEFAULT_GUI_SETTINGS["rowMode"])
    if row_mode not in {"latest", "daily"}:
        row_mode = DEFAULT_GUI_SETTINGS["rowMode"]
    return {"version": 2, "mode": mode, "rowMode": row_mode, "selectedFields": fields or DEFAULT_SELECTED_FIELDS[:]}


def save_settings(data: dict[str, Any]) -> dict[str, Any]:
    fields = data.get("selectedFields")
    if not isinstance(fields, list):
        fields = DEFAULT_SELECTED_FIELDS[:]
    fields = [str(field) for field in fields if str(field) in FIELD_LABELS]
    mode = str(data.get("mode") or DEFAULT_GUI_SETTINGS["mode"])
    if mode not in wechat_stats.ENDPOINTS:
        mode = DEFAULT_GUI_SETTINGS["mode"]
    row_mode = str(data.get("rowMode") or DEFAULT_GUI_SETTINGS["rowMode"])
    if row_mode not in {"latest", "daily"}:
        row_mode = DEFAULT_GUI_SETTINGS["rowMode"]
    settings = {"version": 2, "mode": mode, "rowMode": row_mode, "selectedFields": fields or DEFAULT_SELECTED_FIELDS[:]}
    wechat_stats.repo_path(".gui_settings.json").write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    return settings


def selected_fields_from_payload(payload: dict[str, Any]) -> list[str]:
    raw = payload.get("selectedFields")
    if not isinstance(raw, list):
        return load_settings()["selectedFields"]
    fields = [str(field) for field in raw if str(field) in FIELD_LABELS]
    return fields or load_settings()["selectedFields"]


def default_region_dir() -> str:
    configured_dir = os.environ.get("WECHAT_REGION_SOURCE_DIR", "").strip()
    if configured_dir:
        return configured_dir
    return ""


def default_region_temp_dir() -> Path:
    return wechat_stats.repo_path(".cache", "region-downloads")


def latest_rows_by_article(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = str(row.get("msgid") or row.get("content_url") or row.get("title") or "")
        if not key:
            continue
        current = latest.get(key)
        if current is None or str(row.get("stat_date", "")) >= str(current.get("stat_date", "")):
            latest[key] = row
    return sorted(latest.values(), key=lambda item: (str(item.get("ref_date", "")), str(item.get("title", ""))))


def read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def parse_video_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text.replace("/", "-")[:10]


def clean_text(value: Any) -> str:
    return str(value or "").replace("\r", "\n").strip()


def parse_video_csv(path: Path, start: dt.date | None = None, end: dt.date | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text = read_text_with_fallback(path)
    return parse_video_csv_text(text, start=start, end=end)


def parse_video_csv_text(text: str, start: dt.date | None = None, end: dt.date | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows = list(csv.DictReader(text.splitlines()))
    unified: list[dict[str, Any]] = []
    raw_rows: list[dict[str, Any]] = []
    for raw in rows:
        ref_date = parse_video_date(str(raw.get("发布时间", "")))
        if ref_date:
            try:
                date_value = dt.date.fromisoformat(ref_date)
            except ValueError:
                date_value = None
            if date_value and start and date_value < start:
                continue
            if date_value and end and date_value > end:
                continue
        normalized_raw = {header: clean_text(raw.get(header, "")) for header in VIDEO_RAW_HEADERS}
        raw_rows.append(normalized_raw)
        unified.append(
            {
                "source_type": "视频号",
                "ref_date": ref_date,
                "stat_date": ref_date,
                "title": clean_text(raw.get("视频描述", "")),
                "content_url": clean_text(raw.get("视频ID", "")),
                "read_user": clean_text(raw.get("播放量", "")),
                "like_user": clean_text(raw.get("推荐", "")),
                "share_user": clean_text(raw.get("分享量", "")),
                "zaikan_user": clean_text(raw.get("喜欢", "")),
                "comment_count": clean_text(raw.get("评论量", "")),
                "collection_user": "",
                "read_finish_rate": clean_text(raw.get("完播率", "")),
                "read_avg_activetime": clean_text(raw.get("平均播放时长", "")),
                "read_subscribe_user": clean_text(raw.get("关注量", "")),
            }
        )
    return unified, raw_rows


def combined_headers_for_fields(fields: list[str]) -> list[str]:
    return [COMBINED_LABEL_OVERRIDES.get(field, FIELD_LABELS[field]) for field in fields if field in FIELD_LABELS]


def combined_row_for_fields(row: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    return {COMBINED_LABEL_OVERRIDES.get(field, FIELD_LABELS[field]): row.get(field, "") for field in fields if field in FIELD_LABELS}


NUMERIC_EXCEL_HEADERS = {
    "阅读/播放数", "点赞", "点赞/推荐", "转发", "喜欢", "留言/评论", "收藏", "阅读后关注",
    "播放量", "推荐", "评论量", "分享量", "关注量", "转发聊天和朋友圈", "设为铃声", "设为状态", "设为朋友圈封面",
}


def excel_cell_value(header: str, value: Any) -> Any:
    if header not in NUMERIC_EXCEL_HEADERS or value in (None, ""):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def append_sheet_rows(sheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="13795B", end_color="13795B", fill_type="solid")
    for row in rows:
        sheet.append([excel_cell_value(header, row.get(header, "")) for header in headers])
    sheet.freeze_panes = "A2"
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = max(12, min(48, len(str(header)) + 8))


def write_combined_workbook(path: Path, combined_rows: list[dict[str, Any]], video_raw_rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "运营汇总"
    headers = combined_headers_for_fields(fields)
    append_sheet_rows(summary, headers, [combined_row_for_fields(row, fields) for row in combined_rows])
    video_sheet = workbook.create_sheet("视频号明细")
    append_sheet_rows(video_sheet, VIDEO_RAW_HEADERS, video_raw_rows)
    workbook.save(path)


def sort_combined_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: (str(row.get("ref_date", "")), str(row.get("title", "")), str(row.get("source_type", ""))))


def write_selected_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    target = path
    try:
        file = target.open("w", encoding="utf-8-sig", newline="")
    except PermissionError:
        target = path.with_name(f"{path.stem}_{int(time.time())}{path.suffix}")
        file = target.open("w", encoding="utf-8-sig", newline="")
    with file:
        writer = csv.DictWriter(file, fieldnames=[COMBINED_LABEL_OVERRIDES.get(field, FIELD_LABELS[field]) for field in fields], extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({COMBINED_LABEL_OVERRIDES.get(field, FIELD_LABELS[field]): row.get(field, "") for field in fields})
    return target


def appmsg_analysis_download_url(msgid: str, publish_date: str, token: str, template_url: str = "") -> str:
    safe_msgid = msgid.strip()
    if safe_msgid and "_" not in safe_msgid:
        safe_msgid = f"{safe_msgid}_1"
    if template_url.strip():
        parts = urlsplit(template_url.strip())
        query = parse_qsl(parts.query, keep_blank_values=True)
        new_query: list[tuple[str, str]] = []
        seen = set()
        for key, value in query:
            lower = key.lower()
            if lower in {"msgid", "mid"}:
                value = safe_msgid
            elif lower in {"publish_date", "publishdate", "date"}:
                value = publish_date
            elif lower == "token":
                value = token.strip()
            elif lower == "download":
                value = "1"
            seen.add(lower)
            new_query.append((key, value))
        if "msgid" not in seen:
            new_query.append(("msgid", safe_msgid))
        if "publish_date" not in seen:
            new_query.append(("publish_date", publish_date))
        if "token" not in seen:
            new_query.append(("token", token.strip()))
        if "download" not in seen:
            new_query.append(("download", "1"))
        return urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(new_query), parts.fragment))
    params = {"action": "detailpage", "msgid": safe_msgid, "publish_date": publish_date, "type": "int", "pageVersion": "1", "token": token.strip(), "lang": "zh_CN", "download": "1"}
    return f"https://mp.weixin.qq.com/misc/appmsganalysis?{urlencode(params)}"


def run_export(job_id: str, payload: dict[str, Any]) -> None:
    def log(line: str) -> None:
        with JOBS_LOCK:
            JOBS[job_id]["log"].append(line)

    try:
        wechat_stats.load_dotenv(wechat_stats.repo_path(".env"))
        appid = os.environ.get("WECHAT_APPID", "").strip()
        secret = os.environ.get("WECHAT_APPSECRET", "").strip()
        if not configured():
            raise RuntimeError("请先在 .env 中填写 WECHAT_APPID 和 WECHAT_APPSECRET。")

        start = dt.date.fromisoformat(str(payload.get("start", "")))
        end = dt.date.fromisoformat(str(payload.get("end", "")))
        mode = str(payload.get("mode") or "totaldetail")
        if mode not in wechat_stats.ENDPOINTS:
            raise RuntimeError("请选择有效的数据接口。")

        days = wechat_stats.iter_days(start, end)
        output_dir = resolve_export_dir(payload.get("out"))
        client = wechat_stats.WeChatClient(appid=appid, secret=secret, cache_dir=wechat_stats.repo_path(".cache"))

        all_rows: list[dict[str, Any]] = []
        all_sources: list[dict[str, Any]] = []
        all_jumps: list[dict[str, Any]] = []
        raw_by_day: dict[str, Any] = {}

        log(f"日期范围：{start.isoformat()} 到 {end.isoformat()}")
        log(f"导出目录：{output_dir}")

        for day in days:
            log(f"正在拉取 {day.isoformat()} ...")
            data = client.datacube(mode, day, force_refresh_token=bool(payload.get("forceRefreshToken")))
            raw_by_day[day.isoformat()] = data
            if mode == "totaldetail":
                rows, sources, jumps = wechat_stats.flatten_totaldetail(data)
                all_rows.extend(rows)
                all_sources.extend(sources)
                all_jumps.extend(jumps)
                log(f"{day.isoformat()} 完成：主表 {len(rows)} 行，阅读来源 {len(sources)} 行。")
            else:
                rows = wechat_stats.flatten_legacy(mode, data)
                all_rows.extend(rows)
                log(f"{day.isoformat()} 完成：主表 {len(rows)} 行。")

        stamp = f"{start.isoformat()}_{end.isoformat()}"
        selected_fields = selected_fields_from_payload(payload)
        selected_rows = latest_rows_by_article(all_rows) if payload.get("rowMode") != "daily" else sorted(all_rows, key=lambda item: (str(item.get("ref_date", "")), str(item.get("title", ""))))
        for row in selected_rows:
            row["source_type"] = "公众号"
        selected_path = write_selected_csv(output_dir / f"wechat_articles_selected_{stamp}.csv", selected_rows, selected_fields)
        files = [str(selected_path)]

        with JOBS_LOCK:
            JOBS[job_id].update(
                {
                    "status": "done",
                    "rows": len(selected_rows),
                    "sources": len(all_sources),
                    "jumps": len(all_jumps),
                    "files": files,
                }
            )
            JOBS[job_id]["log"].append("导出完成。")
    except Exception as exc:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["log"].append(f"失败：{exc}")


def run_combined_export(job_id: str, payload: dict[str, Any]) -> None:
    def log(line: str) -> None:
        with JOBS_LOCK:
            JOBS[job_id]["log"].append(line)

    try:
        wechat_stats.load_dotenv(wechat_stats.repo_path(".env"))
        include_articles = bool(payload.get("includeArticles", True))
        appid = os.environ.get("WECHAT_APPID", "").strip()
        secret = os.environ.get("WECHAT_APPSECRET", "").strip()
        if include_articles and not configured():
            raise RuntimeError("请先在密钥设置中填写 AppID 和 AppSecret。")
        start = dt.date.fromisoformat(str(payload.get("start", "")))
        end = dt.date.fromisoformat(str(payload.get("end", "")))
        if start > end:
            raise RuntimeError("开始日期不能晚于结束日期。")
        output_dir = resolve_export_dir(payload.get("out"))
        output_name = str(payload.get("outputName") or "").strip()
        if not output_name:
            output_name = f"wechat_combined_articles_channels_{start.isoformat()}_{end.isoformat()}.xlsx"
        if not output_name.lower().endswith(".xlsx"):
            output_name += ".xlsx"
        output_path = output_dir / output_name
        selected_fields = selected_fields_from_payload(payload)

        article_combined: list[dict[str, Any]] = []
        if include_articles:
            log("正在拉取公众号数据...")
            client = wechat_stats.WeChatClient(appid=appid, secret=secret, cache_dir=wechat_stats.repo_path(".cache"))
            article_rows: list[dict[str, Any]] = []
            for day in wechat_stats.iter_days(start, end):
                data = client.datacube("totaldetail", day, force_refresh_token=bool(payload.get("forceRefreshToken")))
                rows, _, _ = wechat_stats.flatten_totaldetail(data)
                article_rows.extend(rows)
                log(f"公众号 {day.isoformat()}：{len(rows)} 行。")
            for row in latest_rows_by_article(article_rows):
                item = dict(row)
                item["source_type"] = "公众号"
                article_combined.append(item)

        video_files = payload.get("videoFiles") if isinstance(payload.get("videoFiles"), list) else []
        video_csv_texts = payload.get("videoCsvTexts") if isinstance(payload.get("videoCsvTexts"), list) else []
        video_combined: list[dict[str, Any]] = []
        video_raw: list[dict[str, Any]] = []
        for item in video_csv_texts:
            if isinstance(item, dict):
                name = str(item.get("name") or "视频号 CSV")
                text = str(item.get("text") or "")
            else:
                name = "视频号 CSV"
                text = str(item or "")
            if not text.strip():
                continue
            rows, raw_rows = parse_video_csv_text(text, start=start, end=end)
            video_combined.extend(rows)
            video_raw.extend(raw_rows)
            log(f"视频号文件 {name}：{len(rows)} 行。")
        for value in video_files:
            path = Path(str(value))
            if not path.exists():
                log(f"视频号文件不存在，已跳过：{path}")
                continue
            rows, raw_rows = parse_video_csv(path, start=start, end=end)
            video_combined.extend(rows)
            video_raw.extend(raw_rows)
            log(f"视频号文件 {path.name}：{len(rows)} 行。")

        combined_rows = sort_combined_rows(article_combined + video_combined)
        write_combined_workbook(output_path, combined_rows, video_raw, selected_fields)
        log(f"已生成合并表：{output_path}")

        if payload.get("deleteVideoFiles"):
            deleted_video_files = 0
            for value in video_files:
                try:
                    path = Path(str(value))
                    if path.exists() and path.is_file():
                        path.unlink()
                        deleted_video_files += 1
                except Exception:
                    pass
            if deleted_video_files:
                log(f"已清理视频号临时 CSV：{deleted_video_files} 个。")

        with JOBS_LOCK:
            JOBS[job_id].update({"status": "done", "rows": len(combined_rows), "fieldCount": len(combined_rows), "files": [str(output_path)]})
            JOBS[job_id]["log"].append("合并导出完成。")
    except Exception as exc:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["log"].append(f"失败：{exc}")


def run_region_links(job_id: str, payload: dict[str, Any]) -> None:
    try:
        wechat_stats.load_dotenv(wechat_stats.repo_path(".env"))
        appid = os.environ.get("WECHAT_APPID", "").strip()
        secret = os.environ.get("WECHAT_APPSECRET", "").strip()
        if not configured():
            raise RuntimeError("请先在密钥设置中填写 AppID 和 AppSecret。")
        start = dt.date.fromisoformat(str(payload.get("start", "")))
        end = dt.date.fromisoformat(str(payload.get("end", "")))
        token = str(payload.get("mpToken") or "").strip()
        template = str(payload.get("downloadTemplate") or "")
        if not token:
            raise RuntimeError("请先打开公众号后台并识别 token。")
        client = wechat_stats.WeChatClient(appid=appid, secret=secret, cache_dir=wechat_stats.repo_path(".cache"))
        link_rows: list[dict[str, str]] = []
        for day in wechat_stats.iter_days(start, end):
            data = client.datacube("totaldetail", day)
            rows, _, _ = wechat_stats.flatten_totaldetail(data)
            for row in latest_rows_by_article(rows):
                msgid = str(row.get("msgid") or "")
                ref_date = str(row.get("ref_date") or day.isoformat())
                title = str(row.get("title") or "")
                if not msgid:
                    continue
                link_rows.append({"ref_date": ref_date, "title": title, "msgid": msgid, "url": appmsg_analysis_download_url(msgid, ref_date, token, template)})
        with JOBS_LOCK:
            JOBS[job_id].update({"status": "done", "rows": len(link_rows), "fieldCount": len(link_rows), "files": [], "links": link_rows})
            JOBS[job_id]["log"].append(f"已生成 {len(link_rows)} 条地域下载链接。")
    except Exception as exc:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["log"].append(f"失败：{exc}")


def run_region_report(job_id: str, payload: dict[str, Any]) -> None:
    try:
        start = dt.date.fromisoformat(str(payload.get("start", "")))
        end = dt.date.fromisoformat(str(payload.get("end", "")))
        output_dir = resolve_export_dir(payload.get("out"))
        output_name = str(payload.get("outputName") or f"泰绩优地域战报_{start.isoformat()}_{end.isoformat()}.xlsx")
        region_folder = Path(str(payload.get("regionFolder") or default_region_dir()))
        output_path = output_dir / output_name
        result = region_report.build_region_report(region_folder, output_path, start, end)
        with JOBS_LOCK:
            JOBS[job_id].update({"status": "done", "rows": result.get("rows", 0), "fieldCount": result.get("articles", 0), "files": [str(output_path)]})
            JOBS[job_id]["log"].append(f"地域战报完成：{output_path}")
    except Exception as exc:
        with JOBS_LOCK:
            JOBS[job_id]["status"] = "error"
            JOBS[job_id]["log"].append(f"失败：{exc}")


def copy_region_files(payload: dict[str, Any]) -> dict[str, Any]:
    target_dir_value = payload.get("targetFolder") or payload.get("regionFolder") or default_region_dir()
    target_dir = Path(str(target_dir_value))
    target_dir.mkdir(parents=True, exist_ok=True)
    delete_sources = bool(payload.get("deleteSources"))
    copied: list[str] = []
    deleted: list[str] = []
    skipped: list[str] = []
    for value in payload.get("files") or []:
        source = Path(str(value))
        if not source.exists() or not source.is_file():
            skipped.append(str(source))
            continue
        if source.suffix.lower() not in {".xls", ".xlsx"}:
            skipped.append(str(source))
            continue
        target = target_dir / source.name
        try:
            same_file = False
            if source.resolve() != target.resolve():
                for attempt in range(5):
                    try:
                        shutil.copy2(source, target)
                        break
                    except PermissionError:
                        if attempt == 4:
                            raise
                        time.sleep(0.5)
            else:
                same_file = True
            copied.append(str(target))
            if delete_sources and not same_file:
                try:
                    source.unlink()
                    deleted.append(str(source))
                except Exception:
                    skipped.append(str(source))
        except Exception:
            skipped.append(str(source))
    return {"copied": copied, "deleted": deleted, "skipped": skipped, "targetDir": str(target_dir)}


def delete_files(payload: dict[str, Any]) -> dict[str, Any]:
    deleted: list[str] = []
    skipped: list[str] = []
    for value in payload.get("files") or []:
        path = Path(str(value))
        try:
            if path.exists() and path.is_file():
                path.unlink()
                deleted.append(str(path))
            else:
                skipped.append(str(path))
        except Exception:
            skipped.append(str(path))
    for value in payload.get("folders") or []:
        path = Path(str(value))
        try:
            if path.exists() and path.is_dir():
                shutil.rmtree(path)
                deleted.append(str(path))
            else:
                skipped.append(str(path))
        except Exception:
            skipped.append(str(path))
    return {"deleted": deleted, "skipped": skipped}


def pick_folder(initial: str | None = None) -> str:
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askdirectory(initialdir=(initial or ""), title="选择文件夹")
        root.destroy()
        return selected or ""
    except Exception:
        return ""


def save_secrets(payload: dict[str, Any]) -> None:
    env_path = wechat_stats.repo_path(".env")
    existing: dict[str, str] = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "=" in line and not line.lstrip().startswith("#"):
                key, value = line.split("=", 1)
                existing[key.strip()] = value.strip()
    appid = str(payload.get("appid") or existing.get("WECHAT_APPID", "")).strip()
    secret = str(payload.get("appsecret") or existing.get("WECHAT_APPSECRET", "")).strip()
    export_dir = str(payload.get("exportDir") or existing.get("WECHAT_EXPORT_DIR", "exports")).strip() or "exports"
    env_path.write_text(f"WECHAT_APPID={appid}\nWECHAT_APPSECRET={secret}\nWECHAT_EXPORT_DIR={export_dir}\n", encoding="utf-8")
    os.environ["WECHAT_APPID"] = appid
    os.environ["WECHAT_APPSECRET"] = secret
    os.environ["WECHAT_EXPORT_DIR"] = export_dir


class Handler(BaseHTTPRequestHandler):
    protocol_version = "HTTP/1.1"

    def do_GET(self) -> None:
        if self.path == "/" or self.path.startswith("/?"):
            self.send_html(HTML)
            return
        if self.path.startswith("/api/status"):
            yesterday = today_yesterday().isoformat()
            wechat_stats.load_dotenv(wechat_stats.repo_path(".env"))
            self.send_json(
                {
                    "configured": configured(),
                    "appid": os.environ.get("WECHAT_APPID", "").strip(),
                    "secretConfigured": bool(os.environ.get("WECHAT_APPSECRET", "").strip()),
                    "yesterday": yesterday,
                    "monthStart": dt.date.today().replace(day=1).isoformat(),
                    "exportDir": os.environ.get("WECHAT_EXPORT_DIR", "exports"),
                    "regionDir": default_region_dir(),
                    "regionTempDir": str(default_region_temp_dir()),
                    "fieldLabels": FIELD_LABELS,
                    "settings": load_settings(),
                }
            )
            return
        if self.path.startswith("/api/job"):
            job_id = self.query_param("id")
            with JOBS_LOCK:
                job = dict(JOBS.get(job_id, {"status": "missing", "log": ["未找到任务。"]}))
            self.send_json(job)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        if self.path.startswith("/api/export"):
            payload = self.read_json()
            job_id = str(int(time.time() * 1000))
            with JOBS_LOCK:
                JOBS[job_id] = {"status": "running", "log": ["任务已创建。"], "rows": 0, "sources": 0, "jumps": 0, "files": []}
            thread = threading.Thread(target=run_export, args=(job_id, payload), daemon=True)
            thread.start()
            self.send_json({"jobId": job_id})
            return
        if self.path.startswith("/api/open-folder"):
            payload = self.read_json()
            path = resolve_export_dir(payload.get("out"))
            path.mkdir(parents=True, exist_ok=True)
            subprocess.Popen(["explorer.exe", "/n,", str(path)])
            self.send_json({"ok": True})
            return
        if self.path.startswith("/api/pick-folder"):
            payload = self.read_json()
            self.send_json({"path": pick_folder(payload.get("initial"))})
            return
        if self.path.startswith("/api/settings"):
            payload = self.read_json()
            self.send_json({"settings": save_settings(payload)})
            return
        if self.path.startswith("/api/secrets"):
            payload = self.read_json()
            save_secrets(payload)
            self.send_json({"ok": True})
            return
        if self.path.startswith("/api/combined-export"):
            payload = self.read_json()
            job_id = str(int(time.time() * 1000))
            with JOBS_LOCK:
                JOBS[job_id] = {"status": "running", "log": ["合并导出任务已创建。"], "rows": 0, "fieldCount": 0, "files": []}
            thread = threading.Thread(target=run_combined_export, args=(job_id, payload), daemon=True)
            thread.start()
            self.send_json({"jobId": job_id})
            return
        if self.path.startswith("/api/region-links"):
            payload = self.read_json()
            job_id = str(int(time.time() * 1000))
            with JOBS_LOCK:
                JOBS[job_id] = {"status": "running", "log": ["地域下载链接任务已创建。"], "rows": 0, "fieldCount": 0, "files": [], "links": []}
            thread = threading.Thread(target=run_region_links, args=(job_id, payload), daemon=True)
            thread.start()
            self.send_json({"jobId": job_id})
            return
        if self.path.startswith("/api/region-report"):
            payload = self.read_json()
            job_id = str(int(time.time() * 1000))
            with JOBS_LOCK:
                JOBS[job_id] = {"status": "running", "log": ["地域战报任务已创建。"], "rows": 0, "fieldCount": 0, "files": []}
            thread = threading.Thread(target=run_region_report, args=(job_id, payload), daemon=True)
            thread.start()
            self.send_json({"jobId": job_id})
            return
        if self.path.startswith("/api/copy-region-files"):
            try:
                payload = self.read_json()
                self.send_json(copy_region_files(payload))
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=500)
            return
        if self.path.startswith("/api/delete-files"):
            try:
                payload = self.read_json()
                self.send_json(delete_files(payload))
            except Exception as exc:
                self.send_json({"error": str(exc)}, status=500)
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8") if length else "{}"
        return json.loads(raw or "{}")

    def query_param(self, name: str) -> str:
        from urllib.parse import parse_qs, urlparse

        return parse_qs(urlparse(self.path).query).get(name, [""])[0]

    def send_html(self, body: str) -> None:
        data = body.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload: dict[str, Any], status: int = 200) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args: Any) -> None:
        return


def main() -> int:
    parser = argparse.ArgumentParser(description=f"启动{APP_TITLE}图形界面")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--no-open", action="store_true")
    args = parser.parse_args()

    server = ThreadingHTTPServer((HOST, args.port), Handler)
    url = f"http://{HOST}:{args.port}"
    print(f"{APP_TITLE}已启动：{url}")
    if not args.no_open:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("已停止。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
